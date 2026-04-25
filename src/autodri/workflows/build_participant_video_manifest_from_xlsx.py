#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build a video manifest CSV for one participant from the study XLSX sheet.

The output CSV is intentionally minimal and compatible with scripts that only
need a `video` column, such as `assign_dual_roi.py`.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import OrderedDict, defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build participant video manifest from XLSX sheet")
    p.add_argument("--xlsx", required=True, help="Path to the study XLSX file")
    p.add_argument("--sheet", required=True, help="Sheet name, e.g. p6")
    p.add_argument("--videos-root", required=True, help="Root directory that contains the participant videos")
    p.add_argument("--out-csv", required=True, help="Output CSV path")
    p.add_argument(
        "--video-exts",
        default=".mp4,.MP4",
        help="Comma-separated video extensions to scan",
    )
    p.add_argument(
        "--time-tolerance-sec",
        type=int,
        default=180,
        help="Fallback fuzzy-match tolerance on start/end timestamps when exact folder match is absent",
    )
    return p.parse_args()


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", str(text).strip())


def normalize_folder_label(text: str) -> str:
    t = normalize_space(text)
    t = t.replace("，", ".").replace(",", ".")

    m = re.fullmatch(r"(\d{4})(\d{2})(\d{2})\s+(\d{6}-\d{6})", t)
    if m:
        _, mm, dd, tail = m.groups()
        return f"{mm}.{dd} {tail}"

    m = re.fullmatch(r"(\d{1,2})[.](\d{1,2})\s+(\d{6}-\d{6})", t)
    if m:
        mm, dd, tail = m.groups()
        return f"{int(mm):02d}.{int(dd):02d} {tail}"

    return t


def parse_label_key(text: str) -> Optional[Tuple[str, int, int]]:
    t = normalize_space(text)
    m = re.fullmatch(r"(\d{4})(\d{2})(\d{2})\s+(\d{6})-(\d{6})", t)
    if m:
        yyyy, mm, dd, start, end = m.groups()
        return (f"{yyyy}{mm}{dd}", hhmmss_to_sec(start), hhmmss_to_sec(end))

    m = re.fullmatch(r"(\d{1,2})(?:[.,]?)(\d{1,2})\s+(\d{6})-(\d{6})", t)
    if m:
        mm, dd, start, end = m.groups()
        return (f"{int(mm):02d}{int(dd):02d}", hhmmss_to_sec(start), hhmmss_to_sec(end))

    return None


def hhmmss_to_sec(text: str) -> int:
    s = str(text).strip()
    hh = int(s[0:2])
    mm = int(s[2:4])
    ss = int(s[4:6])
    return hh * 3600 + mm * 60 + ss


def read_targets(xlsx_path: Path, sheet_name: str) -> List[str]:
    wb = load_workbook(xlsx_path, read_only=False, data_only=True)
    if sheet_name not in wb.sheetnames:
        # Some files may use P12 instead of p12. Fall back to case-insensitive match.
        sheet_lookup = {name.lower(): name for name in wb.sheetnames}
        key = sheet_name.lower()
        if key not in sheet_lookup:
            raise ValueError(f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}")
        sheet_name = sheet_lookup[key]

    ws = wb[sheet_name]
    out: "OrderedDict[str, None]" = OrderedDict()
    skip_tokens = {"视频文件夹", "0"}

    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        raw = normalize_space(str(v))
        if not raw or raw in skip_tokens:
            continue
        out[raw] = None

    return list(out.keys())


def iter_video_files(videos_root: Path, exts: Sequence[str]) -> Iterable[Path]:
    allow = {e.lower() for e in exts}
    for p in videos_root.rglob("*"):
        if p.is_file() and p.suffix.lower() in allow:
            yield p


def choose_best_video(paths: Sequence[Path]) -> Path:
    def sort_key(p: Path) -> Tuple[int, int, str]:
        size = int(p.stat().st_size)
        # Prefer non-zero, larger files, then stable path ordering.
        return (1 if size > 0 else 0, size, str(p))

    return sorted(paths, key=sort_key, reverse=True)[0]


def build_video_index(videos_root: Path, exts: Sequence[str]) -> Dict[str, Path]:
    by_folder: Dict[str, List[Path]] = defaultdict(list)
    for p in iter_video_files(videos_root, exts):
        by_folder[normalize_folder_label(p.parent.name)].append(p)

    out: Dict[str, Path] = {}
    for folder_key, paths in by_folder.items():
        out[folder_key] = choose_best_video(paths)
    return out


def build_folder_candidates(videos_root: Path, exts: Sequence[str]) -> List[Tuple[str, Path, Optional[Tuple[str, int, int]]]]:
    by_folder: Dict[str, List[Path]] = defaultdict(list)
    for p in iter_video_files(videos_root, exts):
        by_folder[p.parent.name].append(p)

    out: List[Tuple[str, Path, Optional[Tuple[str, int, int]]]] = []
    for folder_name, paths in by_folder.items():
        best = choose_best_video(paths)
        out.append((folder_name, best, parse_label_key(folder_name)))
    return out


def choose_fuzzy_match(
    raw_label: str,
    candidates: Sequence[Tuple[str, Path, Optional[Tuple[str, int, int]]]],
    tolerance_sec: int,
) -> Optional[Tuple[str, Path]]:
    target_key = parse_label_key(raw_label)
    if target_key is None:
        return None

    target_date, target_start, target_end = target_key
    scored: List[Tuple[int, int, str, Path]] = []
    for folder_name, video_path, cand_key in candidates:
        if cand_key is None:
            continue
        cand_date, cand_start, cand_end = cand_key
        if cand_date != target_date:
            continue
        start_diff = abs(cand_start - target_start)
        end_diff = abs(cand_end - target_end)
        if max(start_diff, end_diff) > max(0, int(tolerance_sec)):
            continue
        scored.append((start_diff + end_diff, max(start_diff, end_diff), folder_name, video_path))

    if not scored:
        return None
    _, _, folder_name, video_path = sorted(scored, key=lambda x: (x[0], x[1], x[2]))[0]
    return folder_name, video_path


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    videos_root = Path(args.videos_root)
    out_csv = Path(args.out_csv)

    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    if not videos_root.exists():
        raise FileNotFoundError(videos_root)

    exts = [x.strip() for x in str(args.video_exts).split(",") if x.strip()]
    targets_raw = read_targets(xlsx_path, args.sheet)
    targets_norm = [normalize_folder_label(x) for x in targets_raw]
    video_index = build_video_index(videos_root, exts)
    folder_candidates = build_folder_candidates(videos_root, exts)

    rows = []
    missing = []
    for raw_label, norm_label in zip(targets_raw, targets_norm):
        video_path = video_index.get(norm_label)
        matched_folder = video_path.parent.name if video_path is not None else ""
        if video_path is None:
            fuzzy = choose_fuzzy_match(raw_label, folder_candidates, tolerance_sec=int(args.time_tolerance_sec))
            if fuzzy is not None:
                matched_folder, video_path = fuzzy
        if video_path is None:
            missing.append((raw_label, norm_label))
            continue
        rows.append(
            {
                "sheet_label": raw_label,
                "normalized_label": norm_label,
                "matched_folder": matched_folder,
                "video": str(video_path.resolve()),
                "file_size": str(int(video_path.stat().st_size)),
            }
        )

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["sheet_label", "normalized_label", "matched_folder", "video", "file_size"],
        )
        w.writeheader()
        for row in rows:
            w.writerow(row)

    print(f"Saved: {out_csv}")
    print(f"targets={len(targets_raw)} matched={len(rows)} missing={len(missing)}")
    if missing:
        print("Missing targets:")
        for raw_label, norm_label in missing:
            print(f"  {raw_label} -> {norm_label}")


if __name__ == "__main__":
    main()
