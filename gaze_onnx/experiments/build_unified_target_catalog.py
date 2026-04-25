#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Build a unified target-video/target-segment catalog from project XLSX files.

Inputs:
- main workbook: 自然驾驶_视频标注情况 (3).xlsx
- p11 workbook: P11.xlsx

Outputs:
- unified target video catalog CSV
- unified target segment catalog CSV
- per-participant matched target-video manifests

The parser treats the first column as the source video-folder label and scans all
remaining non-empty cells on the row for time ranges such as:
- 05:31-08:52
- 1:11:04-1:12:43
- 20:00-29：00
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


TIME_RANGE_RE = re.compile(
    r"([0-9]{1,3}[：:][0-9]{2}(?:[：:][0-9]{2})?)\s*[-~—–－到]+\s*([0-9]{1,3}[：:][0-9]{2}(?:[：:][0-9]{2})?)"
)
SKIP_TOKENS = {"", "0", "视频文件夹"}


@dataclass(frozen=True)
class WorkbookSource:
    source_id: str
    xlsx_path: Path
    participant: str
    sheet_name: str


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build unified target catalogs from project XLSX files")
    p.add_argument(
        "--main-xlsx",
        default="自然驾驶_视频标注情况 (3).xlsx",
        help="Main workbook path",
    )
    p.add_argument(
        "--p11-xlsx",
        default="P11.xlsx",
        help="Updated p11 workbook path",
    )
    p.add_argument(
        "--out-videos-csv",
        default="gaze_onnx/experiments/manifests/target_videos.unified.current.csv",
        help="Unified target-video catalog output path",
    )
    p.add_argument(
        "--out-segments-csv",
        default="gaze_onnx/experiments/manifests/target_segments.unified.current.csv",
        help="Unified target-segment catalog output path",
    )
    p.add_argument(
        "--per-participant-dir",
        default="gaze_onnx/experiments/manifests/current",
        help="Output directory for per-participant matched manifests",
    )
    return p.parse_args()


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", str(text).strip())


def normalize_folder_label(text: str) -> str:
    t = normalize_space(text)
    t = t.replace("，", ".").replace(",", ".")

    m = re.fullmatch(r"(\d{4})(\d{2})(\d{2})\s+(\d{6}-\d{6})", t)
    if m:
        _, mm, dd, tail = m.groups()
        return f"{mm}.{dd} {tail}"

    m = re.fullmatch(r"(\d{1,2})[.](\d{1,2})\s+(\d{6}-\d{6})", t)
    if m:
        mm, dd, tail = m.groups()
        return f"{int(mm):02d}.{int(dd):02d} {tail}"

    return t


def hhmmss_to_sec(text: str) -> int:
    s = str(text).strip()
    hh = int(s[0:2])
    mm = int(s[2:4])
    ss = int(s[4:6])
    return hh * 3600 + mm * 60 + ss


def parse_label_key(text: str) -> Optional[Tuple[str, int, int]]:
    t = normalize_space(text)
    m = re.fullmatch(r"(\d{4})(\d{2})(\d{2})\s+(\d{6})-(\d{6})", t)
    if m:
        yyyy, mm, dd, start, end = m.groups()
        return (f"{yyyy}{mm}{dd}", hhmmss_to_sec(start), hhmmss_to_sec(end))

    m = re.fullmatch(r"(\d{1,2})(?:[.,]?)(\d{1,2})\s+(\d{6})-(\d{6})", t)
    if m:
        mm, dd, start, end = m.groups()
        return (f"{int(mm):02d}{int(dd):02d}", hhmmss_to_sec(start), hhmmss_to_sec(end))

    return None


def parse_time_token(token: str) -> Optional[float]:
    t = normalize_space(token).replace("：", ":")
    if not t:
        return None
    parts = t.split(":")
    try:
        vals = [int(x) for x in parts]
    except Exception:
        return None
    if len(vals) == 2:
        mm, ss = vals
        if not (0 <= ss < 60):
            return None
        return float(mm * 60 + ss)
    if len(vals) == 3:
        hh, mm, ss = vals
        if not (0 <= mm < 60 and 0 <= ss < 60):
            return None
        return float(hh * 3600 + mm * 60 + ss)
    return None


def iter_video_files(videos_root: Path) -> Iterable[Path]:
    for p in videos_root.rglob("*"):
        if p.is_file() and p.suffix.lower() == ".mp4":
            yield p


def choose_best_video(paths: Sequence[Path]) -> Path:
    def sort_key(p: Path) -> Tuple[int, int, str]:
        size = int(p.stat().st_size)
        return (1 if size > 0 else 0, size, str(p))

    return sorted(paths, key=sort_key, reverse=True)[0]


def build_video_index(videos_root: Path) -> Dict[str, Path]:
    by_folder: Dict[str, List[Path]] = defaultdict(list)
    for p in iter_video_files(videos_root):
        by_folder[normalize_folder_label(p.parent.name)].append(p)

    out: Dict[str, Path] = {}
    for folder_key, paths in by_folder.items():
        out[folder_key] = choose_best_video(paths)
    return out


def build_folder_candidates(videos_root: Path) -> List[Tuple[str, Path, Optional[Tuple[str, int, int]]]]:
    by_folder: Dict[str, List[Path]] = defaultdict(list)
    for p in iter_video_files(videos_root):
        by_folder[p.parent.name].append(p)

    out: List[Tuple[str, Path, Optional[Tuple[str, int, int]]]] = []
    for folder_name, paths in by_folder.items():
        best = choose_best_video(paths)
        out.append((folder_name, best, parse_label_key(folder_name)))
    return out


def choose_fuzzy_match(
    raw_label: str,
    candidates: Sequence[Tuple[str, Path, Optional[Tuple[str, int, int]]]],
    tolerance_sec: int = 180,
) -> Optional[Tuple[str, Path]]:
    target_key = parse_label_key(raw_label)
    if target_key is None:
        return None

    target_date, target_start, target_end = target_key
    scored: List[Tuple[int, int, str, Path]] = []
    for folder_name, video_path, cand_key in candidates:
        if cand_key is None:
            continue
        cand_date, cand_start, cand_end = cand_key
        if cand_date != target_date:
            continue
        start_diff = abs(cand_start - target_start)
        end_diff = abs(cand_end - target_end)
        if max(start_diff, end_diff) > max(0, int(tolerance_sec)):
            continue
        scored.append((start_diff + end_diff, max(start_diff, end_diff), folder_name, video_path))

    if not scored:
        return None
    _, _, folder_name, video_path = sorted(scored, key=lambda x: (x[0], x[1], x[2]))[0]
    return folder_name, video_path


def resolve_videos_root(participant: str) -> Optional[Path]:
    cand_paths = []
    if participant == "p1":
        cand_paths.append(Path("data/natural_driving_p1"))
    cand_paths.append(Path("data/natural_driving") / participant)
    for p in cand_paths:
        if p.exists():
            return p
    return None


def collect_sources(main_xlsx: Path, p11_xlsx: Path) -> List[WorkbookSource]:
    sources: List[WorkbookSource] = []

    wb = load_workbook(main_xlsx, read_only=False, data_only=True)
    for sheet_name in wb.sheetnames:
        part = sheet_name.strip()
        if part.lower() == "p11":
            continue
        if re.fullmatch(r"p\d+", part.lower()):
            participant = part.lower()
        elif re.fullmatch(r"P\d+", part):
            participant = part.lower()
        else:
            continue
        sources.append(
            WorkbookSource(
                source_id=main_xlsx.name,
                xlsx_path=main_xlsx,
                participant=participant,
                sheet_name=sheet_name,
            )
        )

    wb_p11 = load_workbook(p11_xlsx, read_only=False, data_only=True)
    if wb_p11.sheetnames:
        sources.append(
            WorkbookSource(
                source_id=p11_xlsx.name,
                xlsx_path=p11_xlsx,
                participant="p11",
                sheet_name=wb_p11.sheetnames[0],
            )
        )

    return sources


def row_text_cells(ws, row_idx: int) -> List[str]:
    vals: List[str] = []
    max_col = min(ws.max_column, 12)
    for c in range(2, max_col + 1):
        v = ws.cell(row_idx, c).value
        if v is None:
            continue
        s = normalize_space(str(v))
        if s:
            vals.append(s)
    return vals


def extract_unique_video_labels(ws) -> List[str]:
    out: "OrderedDict[str, None]" = OrderedDict()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        s = normalize_space(str(v))
        if s in SKIP_TOKENS:
            continue
        out[s] = None
    return list(out.keys())


def match_video(label: str, video_index: Dict[str, Path], candidates) -> Tuple[str, str, str]:
    norm = normalize_folder_label(label)
    video_path = video_index.get(norm)
    if video_path is not None:
        return "exact", video_path.parent.name, str(video_path.resolve())
    fuzzy = choose_fuzzy_match(label, candidates)
    if fuzzy is not None:
        folder, video = fuzzy
        return "fuzzy", folder, str(video.resolve())
    return "unmatched", "", ""


def build_catalogs(
    sources: Sequence[WorkbookSource],
) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], Dict[str, List[Dict[str, str]]]]:
    video_rows: List[Dict[str, str]] = []
    segment_rows: List[Dict[str, str]] = []
    participant_manifests: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    segment_uid = 0

    for src in sources:
        wb = load_workbook(src.xlsx_path, read_only=False, data_only=True)
        ws = wb[src.sheet_name]
        videos_root = resolve_videos_root(src.participant)
        video_index = build_video_index(videos_root) if videos_root is not None else {}
        folder_candidates = build_folder_candidates(videos_root) if videos_root is not None else []

        matched_by_label: Dict[str, Tuple[str, str, str]] = {}
        for label in extract_unique_video_labels(ws):
            match_status, matched_folder, video_path = match_video(label, video_index, folder_candidates)
            file_size = ""
            if video_path:
                try:
                    file_size = str(Path(video_path).stat().st_size)
                except Exception:
                    file_size = ""

            row = {
                "participant": src.participant,
                "source_xlsx": src.source_id,
                "source_sheet": src.sheet_name,
                "sheet_label": label,
                "normalized_label": normalize_folder_label(label),
                "videos_root": "" if videos_root is None else str(videos_root.resolve()),
                "matched_folder": matched_folder,
                "video": video_path,
                "file_size": file_size,
                "match_status": match_status,
            }
            video_rows.append(row)
            matched_by_label[label] = (match_status, matched_folder, video_path)
            if video_path:
                participant_manifests[src.participant].append(row)

        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is None:
                continue
            label = normalize_space(str(v))
            if label in SKIP_TOKENS:
                continue

            extra_cells = row_text_cells(ws, r)
            if not extra_cells:
                continue
            joined_text = " | ".join(extra_cells)
            matches = list(TIME_RANGE_RE.finditer(joined_text.replace(" ", "")))
            if not matches:
                continue

            match_status, matched_folder, video_path = matched_by_label.get(label, ("unmatched", "", ""))
            for j, m in enumerate(matches):
                t0 = parse_time_token(m.group(1))
                t1 = parse_time_token(m.group(2))
                if t0 is None or t1 is None or t1 <= t0:
                    continue
                segment_uid += 1
                segment_rows.append(
                    {
                        "segment_uid": f"{src.participant}_seg_{segment_uid:05d}",
                        "participant": src.participant,
                        "source_xlsx": src.source_id,
                        "source_sheet": src.sheet_name,
                        "source_row": str(r),
                        "segment_idx_in_row": str(j),
                        "sheet_label": label,
                        "normalized_label": normalize_folder_label(label),
                        "match_status": match_status,
                        "matched_folder": matched_folder,
                        "video_path": video_path,
                        "time_range_text": f"{m.group(1)}-{m.group(2)}".replace("：", ":"),
                        "start_sec": f"{t0:.3f}",
                        "end_sec": f"{t1:.3f}",
                        "duration_sec": f"{(t1 - t0):.3f}",
                        "row_text": joined_text,
                    }
                )

    video_rows.sort(key=lambda r: (r["participant"], r["normalized_label"], r["sheet_label"]))
    segment_rows.sort(
        key=lambda r: (
            r["participant"],
            r["normalized_label"],
            float(r["start_sec"]),
            int(r["source_row"]),
        )
    )
    for part, rows in participant_manifests.items():
        rows.sort(key=lambda r: (r["normalized_label"], r["sheet_label"], r["video"]))
    return video_rows, segment_rows, participant_manifests


def write_csv(path: Path, rows: List[Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        with path.open("w", encoding="utf-8", newline="") as f:
            csv.writer(f).writerow(["empty"])
        return
    fields = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def main() -> None:
    args = parse_args()
    main_xlsx = Path(args.main_xlsx)
    p11_xlsx = Path(args.p11_xlsx)

    if not main_xlsx.exists():
        raise FileNotFoundError(main_xlsx)
    if not p11_xlsx.exists():
        raise FileNotFoundError(p11_xlsx)

    sources = collect_sources(main_xlsx, p11_xlsx)
    video_rows, segment_rows, participant_manifests = build_catalogs(sources)

    write_csv(Path(args.out_videos_csv), video_rows)
    write_csv(Path(args.out_segments_csv), segment_rows)

    out_dir = Path(args.per_participant_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    for participant, rows in sorted(participant_manifests.items()):
        out_path = out_dir / f"{participant}_target_videos.current.csv"
        manifest_rows = [
            {
                "sheet_label": r["sheet_label"],
                "normalized_label": r["normalized_label"],
                "matched_folder": r["matched_folder"],
                "video": r["video"],
                "file_size": r["file_size"],
                "source_xlsx": r["source_xlsx"],
                "source_sheet": r["source_sheet"],
                "match_status": r["match_status"],
            }
            for r in rows
        ]
        write_csv(out_path, manifest_rows)

    print(f"Sources: {len(sources)}")
    print(f"Unified target videos:   {len(video_rows)} -> {args.out_videos_csv}")
    print(f"Unified target segments: {len(segment_rows)} -> {args.out_segments_csv}")
    for participant in sorted({s.participant for s in sources}):
        video_n = sum(1 for r in video_rows if r["participant"] == participant)
        matched_n = sum(1 for r in video_rows if r["participant"] == participant and r["video"])
        seg_n = sum(1 for r in segment_rows if r["participant"] == participant)
        print(f"{participant}: targets={video_n} matched={matched_n} segments={seg_n}")


if __name__ == "__main__":
    main()
