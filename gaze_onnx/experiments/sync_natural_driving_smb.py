#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Sync Natural Driving pX videos from SMB share without mounting.

Features:
- No CIFS mount required (uses smbclient).
- Incremental sync by file size check (skip already completed files).
- Supports recursive traversal under participant folder.
- Emits final remote/local consistency summary.

Example:
  python gaze_onnx/experiments/sync_natural_driving_smb.py \
    --user nyz \
    --password-env SMB_PASSWORD \
    --participant p1 \
    --out-root data/natural_driving
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

from openpyxl import load_workbook


@dataclass
class RemoteFile:
    rel_path: str
    size: int


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Sync Natural Driving participant videos via SMB")
    p.add_argument("--server", default="10.30.37.3")
    p.add_argument("--share", default="HIS Project Data")
    p.add_argument("--project-root", default="Natural Driving Study - Zhixiong Wang&Jiahao ZHANG/数据【需要拷贝】")
    p.add_argument(
        "--remote-base",
        default="",
        help="Optional full remote base path. Overrides --project-root/--participant/--remote-subdir.",
    )
    p.add_argument("--participant", required=True, help="Participant folder, e.g. p1/p2/p10")
    p.add_argument("--remote-subdir", default="剪辑好的视频", help="Subdir under participant")
    p.add_argument(
        "--local-base",
        default="",
        help="Optional full local base path. Overrides --out-root/participant/remote-subdir.",
    )
    p.add_argument("--user", required=True)
    p.add_argument("--password", default="", help="Plain password (prefer --password-env)")
    p.add_argument("--password-env", default="", help="Read password from env var name")
    p.add_argument("--out-root", default="data/natural_driving")
    p.add_argument("--exts", default=".mp4,.MP4", help="Comma-separated extensions to sync")
    p.add_argument("--target-xlsx", default="", help="Optional XLSX path used to filter target folders")
    p.add_argument("--target-sheet", default="", help="Optional sheet name used with --target-xlsx")
    p.add_argument(
        "--largest-per-folder",
        action="store_true",
        help="Keep only the largest video file under each matched folder",
    )
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--limit", type=int, default=0, help="Debug only: max number of files to transfer")
    return p.parse_args()


def run_cmd(cmd: Sequence[str], check: bool = True) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, check=check, text=True, capture_output=True)


def build_cred_file(user: str, password: str) -> str:
    fd, path = tempfile.mkstemp(prefix="smbcred.", dir="/tmp")
    os.close(fd)
    p = Path(path)
    p.write_text(f"username={user}\npassword={password}\n", encoding="utf-8")
    p.chmod(0o600)
    return str(p)


def normalize_rel(path_str: str) -> str:
    return path_str.replace("\\", "/").strip("/")


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", str(text).strip())


def normalize_folder_label(text: str) -> str:
    t = normalize_space(text)
    t = t.replace("，", ".").replace(",", ".")

    m = re.fullmatch(r"(\d{4})(\d{2})(\d{2})\s+(\d{6}-\d{6})", t)
    if m:
        _, mm, dd, tail = m.groups()
        return f"{mm}.{dd} {tail}"

    m = re.fullmatch(r"(\d{1,2})[.](\d{1,2})\s+(\d{6}-\d{6})", t)
    if m:
        mm, dd, tail = m.groups()
        return f"{int(mm):02d}.{int(dd):02d} {tail}"

    return t


def read_targets(xlsx_path: Path, sheet_name: str) -> List[str]:
    wb = load_workbook(xlsx_path, read_only=False, data_only=True)
    if sheet_name not in wb.sheetnames:
        lookup = {name.lower(): name for name in wb.sheetnames}
        key = sheet_name.lower()
        if key not in lookup:
            raise ValueError(f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}")
        sheet_name = lookup[key]

    ws = wb[sheet_name]
    out: List[str] = []
    seen = set()
    skip_tokens = {"视频文件夹", "0"}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        raw = normalize_space(str(v))
        if not raw or raw in skip_tokens:
            continue
        norm = normalize_folder_label(raw)
        if norm in seen:
            continue
        seen.add(norm)
        out.append(norm)
    return out


def filter_remote_files(
    remote_files: Dict[str, int],
    target_norms: Sequence[str],
    largest_per_folder: bool,
) -> Dict[str, int]:
    if not target_norms:
        return dict(remote_files)

    target_set = set(target_norms)
    by_folder: Dict[str, List[RemoteFile]] = {}
    for rel, size in remote_files.items():
        folder_name = normalize_folder_label(Path(rel).parent.name)
        if folder_name not in target_set:
            continue
        by_folder.setdefault(folder_name, []).append(RemoteFile(rel_path=rel, size=size))

    out: Dict[str, int] = {}
    for items in by_folder.values():
        if largest_per_folder:
            chosen = sorted(items, key=lambda x: (int(x.size), x.rel_path), reverse=True)[0]
            out[chosen.rel_path] = chosen.size
            continue
        for item in items:
            out[item.rel_path] = item.size
    return out


def parse_smb_ls_recursive(raw: str, remote_base: str, allow_exts: Iterable[str]) -> Dict[str, int]:
    allow = {e.lower() for e in allow_exts}
    file_line = re.compile(r"^\s+(.*?)\s+([A-Z]+)\s+([0-9]+)\s+[A-Z][a-z]{2}\s")
    header_prefix = "\\" + remote_base.replace("/", "\\")

    remote: Dict[str, int] = {}
    cur_sub = ""
    for line in raw.splitlines():
        s = line.strip()
        if s.startswith(header_prefix):
            tail = s[len(header_prefix):].lstrip("\\")
            cur_sub = tail.replace("\\", "/")
            continue
        m = file_line.match(line)
        if not m:
            continue
        name, attrs, size_str = m.group(1), m.group(2), m.group(3)
        if "D" in attrs:
            continue
        ext = Path(name).suffix.lower()
        if ext not in allow:
            continue
        rel = f"{cur_sub}/{name}" if cur_sub else name
        rel = normalize_rel(rel)
        remote[rel] = int(size_str)
    return remote


def list_remote_files(server: str, share: str, cred_file: str, remote_base: str, allow_exts: Iterable[str]) -> Dict[str, int]:
    cmd = [
        "smbclient",
        f"//{server}/{share}",
        "-A",
        cred_file,
        "-c",
        f'cd "{remote_base}"; recurse ON; ls',
    ]
    cp = run_cmd(cmd, check=True)
    return parse_smb_ls_recursive(cp.stdout, remote_base=remote_base, allow_exts=allow_exts)


def get_local_files(base: Path, allow_exts: Iterable[str]) -> Dict[str, int]:
    allow = {e.lower() for e in allow_exts}
    out: Dict[str, int] = {}
    if not base.exists():
        return out
    for p in base.rglob("*"):
        if not p.is_file():
            continue
        if p.suffix.lower() not in allow:
            continue
        rel = normalize_rel(str(p.relative_to(base)))
        out[rel] = p.stat().st_size
    return out


def download_one(
    server: str,
    share: str,
    cred_file: str,
    remote_base: str,
    rel_file: str,
    local_base: Path,
    dry_run: bool = False,
) -> None:
    rel_dir = str(Path(rel_file).parent).replace("\\", "/")
    file_name = Path(rel_file).name
    local_dir = local_base / rel_dir
    local_dir.mkdir(parents=True, exist_ok=True)
    if dry_run:
        print(f"[DRY] {rel_file} -> {local_dir}")
        return
    remote_dir = f"{remote_base}/{rel_dir}" if rel_dir not in ("", ".") else remote_base
    cmd = [
        "smbclient",
        f"//{server}/{share}",
        "-A",
        cred_file,
        "-c",
        f'lcd "{local_dir}"; cd "{remote_dir}"; get "{file_name}"',
    ]
    cp = run_cmd(cmd, check=True)
    if cp.stdout.strip():
        print(cp.stdout.strip())


def main() -> None:
    args = parse_args()
    if shutil.which("smbclient") is None:
        raise SystemExit("smbclient not found. Install with: sudo apt-get install -y smbclient")

    password = args.password
    if args.password_env:
        password = os.environ.get(args.password_env, "")
    if not password:
        raise SystemExit("Password is empty. Use --password or --password-env.")

    exts = [x.strip() for x in args.exts.split(",") if x.strip()]
    remote_base = str(args.remote_base).strip()
    if not remote_base:
        remote_base = f"{args.project_root}/{args.participant}/{args.remote_subdir}"

    local_base_arg = str(args.local_base).strip()
    if local_base_arg:
        local_base = Path(local_base_arg)
    else:
        local_base = Path(args.out_root) / args.participant / args.remote_subdir
    local_base.mkdir(parents=True, exist_ok=True)

    cred_file = build_cred_file(args.user, password)
    try:
        print(f"[INFO] remote_base={remote_base}")
        print(f"[INFO] local_base={local_base}")

        remote_files_all = list_remote_files(
            server=args.server,
            share=args.share,
            cred_file=cred_file,
            remote_base=remote_base,
            allow_exts=exts,
        )
        target_norms: List[str] = []
        if args.target_xlsx or args.target_sheet:
            if not (args.target_xlsx and args.target_sheet):
                raise SystemExit("--target-xlsx and --target-sheet must be used together.")
            target_norms = read_targets(Path(args.target_xlsx), args.target_sheet)
        remote_files = filter_remote_files(
            remote_files_all,
            target_norms=target_norms,
            largest_per_folder=bool(args.largest_per_folder),
        )
        local_files = get_local_files(local_base, exts)

        to_transfer: List[RemoteFile] = []
        for rel, rsz in sorted(remote_files.items()):
            lsz = local_files.get(rel)
            if lsz == rsz:
                continue
            to_transfer.append(RemoteFile(rel_path=rel, size=rsz))

        if args.limit > 0:
            to_transfer = to_transfer[: args.limit]

        if target_norms:
            print(f"[INFO] target_folders={len(target_norms)} matched_remote_files={len(remote_files)}")
            print(f"[INFO] remote_files_all={len(remote_files_all)} filtered_remote_files={len(remote_files)}")
        print(f"[INFO] remote_files={len(remote_files)} local_files={len(local_files)} pending={len(to_transfer)}")
        for i, item in enumerate(to_transfer, 1):
            print(f"[{i}/{len(to_transfer)}] {item.rel_path} ({item.size} bytes)")
            download_one(
                server=args.server,
                share=args.share,
                cred_file=cred_file,
                remote_base=remote_base,
                rel_file=item.rel_path,
                local_base=local_base,
                dry_run=args.dry_run,
            )

        # Final verification
        local_files_after = get_local_files(local_base, exts)
        missing = [k for k in remote_files if k not in local_files_after]
        mismatch = [k for k in remote_files if k in local_files_after and remote_files[k] != local_files_after[k]]
        extra = [k for k in local_files_after if k not in remote_files]

        print("\n=== Sync Summary ===")
        print(f"remote_count={len(remote_files)}")
        print(f"local_count={len(local_files_after)}")
        print(f"missing={len(missing)} mismatch={len(mismatch)} extra={len(extra)}")
        print(f"remote_total_bytes={sum(remote_files.values())}")
        print(f"local_total_bytes={sum(local_files_after.values())}")
        if missing:
            print("[WARN] missing examples:")
            for x in missing[:10]:
                print(f"  {x}")
        if mismatch:
            print("[WARN] mismatch examples:")
            for x in mismatch[:10]:
                print(f"  {x} remote={remote_files[x]} local={local_files_after.get(x)}")
    finally:
        try:
            os.remove(cred_file)
        except OSError:
            pass


if __name__ == "__main__":
    main()
